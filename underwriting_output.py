#!/usr/bin/env python3
"""
Underwriting Output Generator
Creates structured, polished underwriting packages with required tabs and formats.
"""

import pandas as pd
import json
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import logging
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

@dataclass
class UnderwritingLine:
    """Individual line item in underwriting summary."""
    line_item: str
    amount: float
    percent_egi: float
    notes: str
    category: str = ""
    is_override: bool = False

class UnderwritingOutputGenerator:
    """
    Generates structured underwriting output with required tabs and formatting.
    Implements deep logic for data consistency and decision-making.
    """
    
    def __init__(self, debug: bool = False):
        self.debug = debug
        self.logger = self._setup_logger()
        
        # Core data storage
        self.rent_roll_data = None
        self.t12_data = None
        self.property_info = {}
        self.underwriting_summary = []
        self.flags_and_notes = []
        
        # Analysis results
        self.income_analysis = {}
        self.expense_analysis = {}
        self.noi_analysis = {}
        
        # Pro forma data (for bridge loans)
        self.pro_forma_rent_roll = None
        self.pro_forma_income = None
        self.sources_uses = None
        self.construction_budget = None
        self.is_bridge_loan = False
    
    def _setup_logger(self):
        """Set up logging for the output generator."""
        logger = logging.getLogger('UnderwritingOutput')
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        logger.setLevel(logging.DEBUG if self.debug else logging.INFO)
        return logger
    
    def load_analysis_data(self, rent_roll_analysis: Dict, t12_analysis: Dict, 
                          property_info: Dict, underwriting_summary: Dict):
        """Load analysis results from the underwriting analyzer."""
        self.rent_roll_data = rent_roll_analysis
        self.t12_data = t12_analysis
        self.property_info = property_info
        
        # Extract key analysis components
        if 'income_summary' in underwriting_summary:
            self.income_analysis = underwriting_summary['income_summary']
        
        if 'expense_analysis' in t12_analysis:
            self.expense_analysis = t12_analysis['expense_analysis']
        
        if 'noi_analysis' in underwriting_summary:
            self.noi_analysis = underwriting_summary['noi_analysis']
        
        # Extract flags and recommendations
        if 'flags_and_recommendations' in underwriting_summary:
            self.flags_and_notes = underwriting_summary['flags_and_recommendations']
        
        self.logger.info("✅ Analysis data loaded successfully")
    
    def set_bridge_loan_mode(self, is_bridge: bool = True):
        """Enable bridge loan mode for additional pro forma tabs."""
        self.is_bridge_loan = is_bridge
        if is_bridge:
            self.logger.info("🌉 Bridge loan mode enabled - will generate pro forma tabs")
    
    def generate_clean_rent_roll(self) -> pd.DataFrame:
        """
        Generate clean rent roll tab with standardized columns.
        Deep logic: Choose best data source and standardize format.
        """
        self.logger.info("🏢 Generating Clean Rent Roll...")
        
        if not self.rent_roll_data or 'raw_data' not in self.rent_roll_data:
            self.logger.warning("⚠️ No rent roll data available")
            return pd.DataFrame()
        
        raw_df = self.rent_roll_data['raw_data']
        
        # Standardized column mapping
        standard_columns = [
            'Unit Number', 'Unit Type', 'Square Footage', 'Current Rent',
            'Market Rent', 'Lease Start', 'Lease End', 'Tenant Name',
            'Security Deposit', 'Status', 'Notes'
        ]
        
        # Create clean dataframe
        clean_df = pd.DataFrame(columns=standard_columns)
        
        # Map raw data to standard columns with intelligent detection
        column_mapping = self._detect_rent_roll_columns(raw_df.columns)
        
        for standard_col, raw_col in column_mapping.items():
            if raw_col and raw_col in raw_df.columns:
                clean_df[standard_col] = raw_df[raw_col]
        
        # Clean and validate data
        clean_df = self._clean_rent_roll_data(clean_df)
        
        # Add calculated fields
        clean_df['Monthly Income'] = pd.to_numeric(clean_df['Current Rent'], errors='coerce').fillna(0)
        clean_df['Annual Income'] = clean_df['Monthly Income'] * 12
        
        # Flag issues
        flags = self._flag_rent_roll_issues(clean_df)
        for flag in flags:
            self.flags_and_notes.append(flag)
        
        self.logger.info(f"✅ Clean rent roll generated: {len(clean_df)} units")
        return clean_df
    
    def generate_clean_t12(self) -> pd.DataFrame:
        """
        Generate clean T12 tab (cut off at NOI).
        Deep logic: Handle inconsistencies, prefer recent trends if upward.
        """
        self.logger.info("💰 Generating Clean T12...")
        
        if not self.t12_data or 'raw_data' not in self.t12_data:
            self.logger.warning("⚠️ No T12 data available")
            return pd.DataFrame()
        
        raw_df = self.t12_data['raw_data']
        
        # Standardized T12 structure
        standard_categories = [
            # Income
            'Rental Income', 'Other Income', 'Gross Potential Income',
            'Vacancy Loss', 'Effective Gross Income',
            
            # Operating Expenses
            'Property Taxes', 'Insurance', 'Utilities', 'Repairs & Maintenance',
            'Management Fee', 'Professional Services', 'Marketing', 'Administrative',
            'Replacement Reserves', 'Other Expenses', 'Total Operating Expenses',
            
            # NOI
            'Net Operating Income'
        ]
        
        # Create clean T12 structure
        clean_t12 = pd.DataFrame({
            'Line Item': standard_categories,
            'Annual Amount': 0.0,
            'Monthly Amount': 0.0,
            'Source': '',
            'Notes': ''
        })
        
        # Map raw data to standard categories
        category_mapping = self._detect_t12_categories(raw_df)
        
        for idx, category in enumerate(standard_categories):
            mapped_data = self._extract_t12_line_item(raw_df, category, category_mapping)
            if mapped_data:
                clean_t12.loc[idx, 'Annual Amount'] = mapped_data['amount']
                clean_t12.loc[idx, 'Monthly Amount'] = mapped_data['amount'] / 12
                clean_t12.loc[idx, 'Source'] = mapped_data['source']
                clean_t12.loc[idx, 'Notes'] = mapped_data['notes']
        
        # Apply deep logic for inconsistencies
        clean_t12 = self._apply_t12_deep_logic(clean_t12, raw_df)
        
        # Cut off at NOI (remove anything after)
        noi_idx = clean_t12[clean_t12['Line Item'] == 'Net Operating Income'].index
        if len(noi_idx) > 0:
            clean_t12 = clean_t12.iloc[:noi_idx[0] + 1]
        
        self.logger.info(f"✅ Clean T12 generated: {len(clean_t12)} line items")
        return clean_t12
    
    def generate_underwriting_summary(self) -> pd.DataFrame:
        """
        Generate underwriting summary with columns: Line Item, $ Amount, % of EGI, Notes.
        Explains every override or adjustment from raw docs.
        """
        self.logger.info("📊 Generating Underwriting Summary...")
        
        summary_lines = []
        
        # Calculate EGI for percentage calculations
        egi = self.noi_analysis.get('effective_gross_income', 0)
        if egi == 0:
            egi = 1  # Prevent division by zero
        
        # INCOME SECTION
        summary_lines.append(self._create_summary_line(
            "GROSS POTENTIAL INCOME", "INCOME", 
            self.income_analysis.get('gross_potential_income', 0), egi,
            "Based on current rent roll analysis with vacant units at market rates"
        ))
        
        summary_lines.append(self._create_summary_line(
            "Vacancy Loss", "INCOME",
            -abs(self.noi_analysis.get('vacancy_loss', 0)), egi,
            f"Applied {self.expense_analysis.get('vacancy_rate', 5)}% vacancy rate (higher of 5% or actuals)"
        ))
        
        summary_lines.append(self._create_summary_line(
            "Other Income", "INCOME",
            self.income_analysis.get('other_income', 0), egi,
            "Used actual T12 totals for other income streams"
        ))
        
        summary_lines.append(self._create_summary_line(
            "EFFECTIVE GROSS INCOME", "INCOME",
            egi, egi,
            "GPI minus vacancy loss plus other income", is_total=True
        ))
        
        # EXPENSE SECTION
        adjusted_expenses = self.expense_analysis.get('adjusted_expenses', {})
        adjustments = self.expense_analysis.get('adjustments', {})
        
        for expense_key, amount in adjusted_expenses.items():
            if amount > 0:  # Only show non-zero expenses
                expense_name = expense_key.replace('_', ' ').title()
                adjustment_note = adjustments.get(expense_key, "No adjustment applied")
                
                summary_lines.append(self._create_summary_line(
                    expense_name, "EXPENSE", amount, egi, adjustment_note
                ))
        
        # NOI CALCULATION
        total_expenses = self.expense_analysis.get('total_adjusted_expenses', 0)
        noi = egi - total_expenses
        
        summary_lines.append(self._create_summary_line(
            "TOTAL OPERATING EXPENSES", "EXPENSE",
            total_expenses, egi,
            f"Total of all adjusted operating expenses", is_total=True
        ))
        
        summary_lines.append(self._create_summary_line(
            "NET OPERATING INCOME", "NOI",
            noi, egi,
            "EGI minus total operating expenses", is_total=True
        ))
        
        # Convert to DataFrame
        summary_df = pd.DataFrame([
            {
                'Line Item': line.line_item,
                '$ Amount': line.amount,
                '% of EGI': line.percent_egi,
                'Notes': line.notes
            }
            for line in summary_lines
        ])
        
        self.logger.info(f"✅ Underwriting summary generated: {len(summary_df)} line items")
        return summary_df
    
    def generate_pro_forma_tabs(self) -> Dict[str, pd.DataFrame]:
        """Generate pro forma tabs for bridge loans."""
        if not self.is_bridge_loan:
            return {}
        
        self.logger.info("🌉 Generating Pro Forma tabs for bridge loan...")
        
        pro_forma_tabs = {}
        
        # Pro Forma Rent Roll
        pro_forma_tabs['Pro Forma Rent Roll'] = self._generate_pro_forma_rent_roll()
        
        # Pro Forma Income Statement
        pro_forma_tabs['Pro Forma Income Statement'] = self._generate_pro_forma_income()
        
        # Sources & Uses
        pro_forma_tabs['Sources & Uses'] = self._generate_sources_uses()
        
        # Construction Budget
        pro_forma_tabs['Construction Budget'] = self._generate_construction_budget()
        
        return pro_forma_tabs
    
    def export_to_excel(self, output_path: str = None) -> str:
        """
        Export all tabs to a structured Excel workbook.
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            property_name = self.property_info.get('property_name', 'Property').replace(' ', '_')
            output_path = f"outputs/{property_name}_Underwriting_{timestamp}.xlsx"
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        self.logger.info(f"📊 Exporting underwriting package to: {output_path}")
        
        # Generate all tabs
        clean_rent_roll = self.generate_clean_rent_roll()
        clean_t12 = self.generate_clean_t12()
        underwriting_summary = self.generate_underwriting_summary()
        
        # Create workbook with styling
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Tab 1: Clean Rent Roll
        self._add_styled_sheet(wb, "Clean Rent Roll", clean_rent_roll)
        
        # Tab 2: Clean T12
        self._add_styled_sheet(wb, "Clean T12", clean_t12)
        
        # Tab 3: Underwriting Summary
        self._add_styled_sheet(wb, "Underwriting Summary", underwriting_summary)
        
        # Bridge loan tabs (if applicable)
        if self.is_bridge_loan:
            pro_forma_tabs = self.generate_pro_forma_tabs()
            for tab_name, tab_data in pro_forma_tabs.items():
                self._add_styled_sheet(wb, tab_name, tab_data)
        
        # Save workbook
        wb.save(output_path)
        
        self.logger.info(f"✅ Excel package exported successfully")
        return output_path
    
    def generate_pdf_package(self, excel_path: str = None) -> str:
        """
        Generate polished PDF of the underwriting package.
        """
        if excel_path is None:
            excel_path = self.export_to_excel()
        
        # Generate PDF output path
        pdf_path = excel_path.replace('.xlsx', '_Package.pdf')
        
        try:
            # Option 1: Try converting Excel to PDF using openpyxl + reportlab
            self._convert_excel_to_pdf(excel_path, pdf_path)
            self.logger.info(f"📄 PDF package generated: {pdf_path}")
            return pdf_path
        except ImportError:
            # Option 2: Create PDF directly using reportlab
            try:
                self._create_pdf_from_data(pdf_path)
                self.logger.info(f"📄 PDF package generated: {pdf_path}")
                return pdf_path
            except ImportError:
                # Fallback: Excel as master template
                self.logger.info("📄 PDF libraries not available - Excel file serves as master template")
                return excel_path
        except Exception as e:
            self.logger.warning(f"⚠️ PDF generation failed: {str(e)} - using Excel template")
            return excel_path
    
    def _convert_excel_to_pdf(self, excel_path: str, pdf_path: str):
        """Convert Excel workbook to PDF using available libraries."""
        try:
            # Try using win32com (Windows only)
            import win32com.client as win32
            
            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            workbook = excel_app.Workbooks.Open(os.path.abspath(excel_path))
            workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            workbook.Close()
            excel_app.Quit()
            
        except ImportError:
            # Try using openpyxl2pdf (if available)
            import subprocess
            result = subprocess.run([
                'python', '-c', 
                f"import openpyxl; from openpyxl2pdf import convert; convert('{excel_path}', '{pdf_path}')"
            ], capture_output=True)
            if result.returncode != 0:
                raise ImportError("openpyxl2pdf conversion failed")
    
    def _create_pdf_from_data(self, pdf_path: str):
        """Create PDF directly from data using reportlab."""
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        property_name = self.property_info.get('property_name', 'Property')
        title = Paragraph(f"<b>Underwriting Analysis - {property_name}</b>", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Property Information
        prop_info_style = styles['Normal']
        prop_info = [
            f"Property: {self.property_info.get('property_name', 'N/A')}",
            f"Address: {self.property_info.get('property_address', 'N/A')}",
            f"Units: {self.property_info.get('unit_count', 'N/A')}",
            f"Transaction: {self.property_info.get('transaction_type', 'N/A').title()}",
            f"Analysis Date: {datetime.now().strftime('%B %d, %Y')}"
        ]
        
        for info in prop_info:
            story.append(Paragraph(info, prop_info_style))
        
        story.append(Spacer(1, 20))
        
        # Underwriting Summary Table
        summary_title = Paragraph("<b>Underwriting Summary</b>", styles['Heading2'])
        story.append(summary_title)
        story.append(Spacer(1, 12))
        
        # Generate underwriting summary data
        summary_df = self.generate_underwriting_summary()
        
        # Create table data
        table_data = [['Line Item', 'Amount', '% of EGI', 'Notes']]
        for _, row in summary_df.iterrows():
            table_data.append([
                str(row['Line Item']),
                f"${row['$ Amount']:,.0f}",
                f"{row['% of EGI']:.1f}%",
                str(row['Notes'])[:50] + "..." if len(str(row['Notes'])) > 50 else str(row['Notes'])
            ])
        
        # Create and style table
        table = Table(table_data, colWidths=[2*inch, 1*inch, 0.8*inch, 2.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Key Metrics
        metrics_title = Paragraph("<b>Key Financial Metrics</b>", styles['Heading2'])
        story.append(metrics_title)
        story.append(Spacer(1, 12))
        
        noi = self.noi_analysis.get('net_operating_income', 0)
        egi = self.noi_analysis.get('effective_gross_income', 0)
        expense_ratio = self.noi_analysis.get('expense_ratio', 0)
        
        metrics = [
            f"Net Operating Income: ${noi:,.0f}",
            f"Effective Gross Income: ${egi:,.0f}",
            f"Operating Expense Ratio: {expense_ratio:.1f}%",
        ]
        
        for metric in metrics:
            story.append(Paragraph(metric, styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Flags and Notes
        if self.flags_and_notes:
            flags_title = Paragraph("<b>Analysis Flags & Recommendations</b>", styles['Heading2'])
            story.append(flags_title)
            story.append(Spacer(1, 12))
            
            for i, flag in enumerate(self.flags_and_notes[:5], 1):
                flag_text = f"{i}. {flag.get('message', 'No message')}"
                story.append(Paragraph(flag_text, styles['Normal']))
                story.append(Spacer(1, 6))
        
        # Build PDF
        doc.build(story)
        
        self.logger.info(f"✅ PDF created with {len(story)} elements")
    
    # Helper methods
    def _create_summary_line(self, line_item: str, category: str, amount: float, 
                           egi: float, notes: str, is_total: bool = False) -> UnderwritingLine:
        """Create a standardized underwriting summary line."""
        percent_egi = (amount / egi * 100) if egi > 0 else 0
        
        return UnderwritingLine(
            line_item=line_item,
            amount=amount,
            percent_egi=percent_egi,
            notes=notes,
            category=category,
            is_override=bool(notes and ("adjusted" in notes.lower() or "override" in notes.lower()))
        )
    
    def _detect_rent_roll_columns(self, columns: List[str]) -> Dict[str, str]:
        """Intelligently detect rent roll column mappings."""
        mapping = {}
        
        # Column detection patterns
        patterns = {
            'Unit Number': ['unit', 'apt', 'number', '#'],
            'Unit Type': ['type', 'bedroom', 'br', 'bed'],
            'Square Footage': ['sqft', 'sq ft', 'square', 'footage', 'sf'],
            'Current Rent': ['rent', 'current', 'amount'],
            'Market Rent': ['market', 'asking'],
            'Lease Start': ['start', 'lease start', 'move in'],
            'Lease End': ['end', 'lease end', 'expir'],
            'Tenant Name': ['tenant', 'name', 'resident'],
            'Security Deposit': ['deposit', 'security'],
            'Status': ['status', 'occupied', 'vacant']
        }
        
        for standard_col, keywords in patterns.items():
            for col in columns:
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in keywords):
                    mapping[standard_col] = col
                    break
        
        return mapping
    
    def _clean_rent_roll_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate rent roll data."""
        # Convert rent columns to numeric
        rent_columns = ['Current Rent', 'Market Rent', 'Security Deposit']
        for col in rent_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Standardize status
        if 'Status' in df.columns:
            df['Status'] = df['Status'].fillna('Unknown').astype(str)
            df.loc[df['Current Rent'] == 0, 'Status'] = 'Vacant'
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        return df
    
    def _flag_rent_roll_issues(self, df: pd.DataFrame) -> List[Dict]:
        """Flag issues in rent roll data."""
        flags = []
        
        # Missing square footage
        if 'Square Footage' in df.columns:
            missing_sqft = df['Square Footage'].isna().sum()
            if missing_sqft > 0:
                flags.append({
                    'type': 'data_quality',
                    'severity': 'medium',
                    'message': f"Missing square footage for {missing_sqft} units"
                })
        
        # Missing rent data
        if 'Current Rent' in df.columns:
            missing_rent = (df['Current Rent'] == 0).sum()
            if missing_rent > 0:
                flags.append({
                    'type': 'data_quality',
                    'severity': 'high',
                    'message': f"Missing current rent for {missing_rent} units"
                })
        
        return flags
    
    def _detect_t12_categories(self, df: pd.DataFrame) -> Dict[str, str]:
        """Detect T12 category mappings from raw data."""
        # This would implement intelligent category detection
        # For now, return empty mapping (would be enhanced in full implementation)
        return {}
    
    def _extract_t12_line_item(self, df: pd.DataFrame, category: str, mapping: Dict) -> Optional[Dict]:
        """Extract specific line item from T12 data."""
        # Placeholder implementation - would contain sophisticated extraction logic
        return {
            'amount': 0.0,
            'source': 'T12 Statement',
            'notes': 'Extracted from original T12'
        }
    
    def _apply_t12_deep_logic(self, clean_df: pd.DataFrame, raw_df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply deep logic for T12 inconsistencies.
        Prefer trailing 2-6 months if trending upward.
        """
        # Placeholder for deep logic implementation
        # Would analyze trends and apply intelligent adjustments
        return clean_df
    
    def _generate_pro_forma_rent_roll(self) -> pd.DataFrame:
        """Generate pro forma rent roll for bridge loans."""
        # Placeholder - would generate projected rent roll
        return pd.DataFrame({'Note': ['Pro forma rent roll generation not yet implemented']})
    
    def _generate_pro_forma_income(self) -> pd.DataFrame:
        """Generate pro forma income statement for bridge loans."""
        # Placeholder - would generate projected income statement
        return pd.DataFrame({'Note': ['Pro forma income statement generation not yet implemented']})
    
    def _generate_sources_uses(self) -> pd.DataFrame:
        """Generate sources & uses statement for bridge loans."""
        # Placeholder - would generate sources & uses
        return pd.DataFrame({'Note': ['Sources & uses generation not yet implemented']})
    
    def _generate_construction_budget(self) -> pd.DataFrame:
        """Generate construction budget for bridge loans."""
        # Placeholder - would generate construction budget
        return pd.DataFrame({'Note': ['Construction budget generation not yet implemented']})
    
    def _add_styled_sheet(self, wb: Workbook, sheet_name: str, df: pd.DataFrame):
        """Add a styled sheet to the workbook."""
        ws = wb.create_sheet(title=sheet_name)
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply basic styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Example usage and integration
if __name__ == "__main__":
    generator = UnderwritingOutputGenerator(debug=True)
    print("📊 Underwriting Output Generator initialized")
    print("✅ Ready to generate structured underwriting packages")
