#!/usr/bin/env python3
"""
FastAPI Underwriting Application - Simplified Demo Version
Modern web interface for real estate underwriting with dynamic uploads and progress tracking.
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import os
import uuid
import json
import asyncio
import shutil
from datetime import datetime, timedelta
import logging

# FastAPI app initialization
app = FastAPI(
    title="Real Estate Underwriting AI",
    description="Professional underwriting analysis with AI-powered document processing",
    version="1.0.0"
)

# CORS middleware for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# In-memory storage for processing status (in production, use Redis/database)
processing_sessions = {}

# Models for API requests/responses
class ProcessingStatus(BaseModel):
    session_id: str
    status: str  # "waiting", "processing", "completed", "error"
    current_step: int
    total_steps: int
    step_name: str
    progress_percentage: float
    message: str
    results: Optional[Dict[str, Any]] = None
    error_message: Optional[str] = None

class PropertyInfo(BaseModel):
    property_name: str
    property_address: str
    transaction_type: str = "refinance"
    is_bridge_loan: bool = False

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the main application page."""
    return FileResponse("templates/index.html")

@app.post("/api/upload")
async def upload_documents(
    background_tasks: BackgroundTasks,
    property_name: str = Form(...),
    property_address: str = Form(...),
    transaction_type: str = Form("refinance"),
    is_bridge_loan: bool = Form(False),
    files: List[UploadFile] = File(...),
    file_types: List[str] = Form([])
):
    """
    Upload documents and start processing in background.
    Returns session ID for tracking progress.
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    
    # Generate unique session ID
    session_id = str(uuid.uuid4())
    
    # Create session directory
    session_dir = f"uploads/{session_id}"
    os.makedirs(session_dir, exist_ok=True)
    
    # Save uploaded files with type information
    uploaded_files = []
    file_type_mapping = {}
    
    for i, file in enumerate(files):
        if file.filename:
            # Get file type (rent_roll, t12, or additional)
            file_type = file_types[i] if i < len(file_types) else 'additional'
            
            # Create type-specific directory
            type_dir = os.path.join(session_dir, file_type)
            os.makedirs(type_dir, exist_ok=True)
            
            file_path = os.path.join(type_dir, file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            uploaded_files.append(file_path)
            file_type_mapping[file_path] = file_type
    
    # Count files by type
    rent_roll_count = sum(1 for ft in file_types if ft == 'rent_roll')
    t12_count = sum(1 for ft in file_types if ft == 't12')
    additional_count = sum(1 for ft in file_types if ft == 'additional')
    
    # Initialize processing status
    processing_sessions[session_id] = ProcessingStatus(
        session_id=session_id,
        status="waiting",
        current_step=0,
        total_steps=7,
        step_name="Initializing...",
        progress_percentage=0.0,
        message=f"Preparing to process {len(uploaded_files)} documents"
    )
    
    # Property information
    property_info = PropertyInfo(
        property_name=property_name,
        property_address=property_address,
        transaction_type=transaction_type,
        is_bridge_loan=is_bridge_loan
    )
    
    # Start background processing
    background_tasks.add_task(
        process_documents_background,
        session_id,
        uploaded_files,
        file_type_mapping,
        property_info
    )
    
    return {
        "session_id": session_id, 
        "message": "Processing started", 
        "files_uploaded": len(uploaded_files),
        "file_breakdown": {
            "rent_roll": rent_roll_count,
            "t12": t12_count,
            "additional": additional_count
        }
    }

@app.get("/api/status/{session_id}")
async def get_processing_status(session_id: str):
    """Get current processing status for a session."""
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return processing_sessions[session_id]

@app.get("/api/results/{session_id}")
async def get_results(session_id: str):
    """Get final results for completed session."""
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_sessions[session_id]
    if session.status != "completed":
        raise HTTPException(status_code=400, detail="Processing not completed")
    
    return session.results

@app.get("/api/download/{session_id}/{file_type}")
async def download_file(session_id: str, file_type: str):
    """Download generated files (excel or pdf)."""
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_sessions[session_id]
    if session.status != "completed" or not session.results:
        raise HTTPException(status_code=400, detail="No files available")
    
    if file_type == "excel":
        file_path = session.results.get("excel_path")
    elif file_type == "pdf":
        file_path = session.results.get("pdf_path")
    else:
        raise HTTPException(status_code=400, detail="Invalid file type")
    
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path,
        media_type="application/octet-stream",
        filename=os.path.basename(file_path)
    )

async def process_documents_background(
    session_id: str,
    uploaded_files: List[str],
    file_type_mapping: Dict[str, str],
    property_info: PropertyInfo
):
    """
    Background task for processing documents with progress updates.
    This is a demo version that simulates the full processing workflow.
    """
    try:
        session = processing_sessions[session_id]
        session.status = "processing"
        
        # Categorize files by type
        rent_roll_files = [f for f, t in file_type_mapping.items() if t == 'rent_roll']
        t12_files = [f for f, t in file_type_mapping.items() if t == 't12']
        additional_files = [f for f, t in file_type_mapping.items() if t == 'additional']
        
        # Step 1: Document Processing
        update_progress(session_id, 1, "Document Processing", 
                       f"Processing {len(rent_roll_files)} rent roll, {len(t12_files)} T12, and {len(additional_files)} additional documents...")
        await asyncio.sleep(2)
        
        # Step 2: Data Analysis Setup
        update_progress(session_id, 2, "Data Analysis Setup", "Initializing underwriting analyzer...")
        await asyncio.sleep(1.5)
        
        # Step 3: Rent Roll Analysis
        if rent_roll_files:
            update_progress(session_id, 3, "Rent Roll Analysis", f"Analyzing {len(rent_roll_files)} rent roll documents...")
        else:
            update_progress(session_id, 3, "Rent Roll Analysis", "No rent roll files - using property assumptions...")
        await asyncio.sleep(2)
        
        # Step 4: T12 Analysis
        if t12_files:
            update_progress(session_id, 4, "T12 Analysis", f"Processing {len(t12_files)} operating statements...")
        else:
            update_progress(session_id, 4, "T12 Analysis", "No T12 files - using market assumptions...")
        await asyncio.sleep(2)
        
        # Step 5: Underwriting Summary
        update_progress(session_id, 5, "Underwriting Summary", "Generating comprehensive analysis...")
        await asyncio.sleep(1.5)
        
        # Step 6: Excel Generation
        update_progress(session_id, 6, "Excel Generation", "Creating professional underwriting package...")
        await asyncio.sleep(2)
        
        # Calculate estimated units and basic metrics early for use throughout
        base_rent_per_unit = 1200 if "apartment" in property_info.property_name.lower() else 1500
        estimated_units = max(50, len(uploaded_files) * 15)  # Estimate units from file complexity
        
        # Calculate quality score based on available files
        quality_score = 0
        if rent_roll_files: quality_score += 40
        if t12_files: quality_score += 40
        if additional_files: quality_score += 20
        
        # Calculate all financial metrics early for Excel and PDF generation
        if rent_roll_files:
            # If rent roll provided, assume higher accuracy
            gross_potential_income = base_rent_per_unit * estimated_units * 12
            vacancy_factor = 0.05 if t12_files else 0.08  # Lower vacancy if T12 confirms
        else:
            # Without rent roll, use conservative estimates
            gross_potential_income = base_rent_per_unit * estimated_units * 12 * 0.85
            vacancy_factor = 0.10
        
        effective_gross_income = gross_potential_income * (1 - vacancy_factor)
        
        # Operating expenses based on property type and location
        expense_ratio = 0.35 if property_info.is_bridge_loan else 0.32  # Higher for bridge loans
        if "luxury" in property_info.property_name.lower() or "premium" in property_info.property_name.lower():
            expense_ratio -= 0.03  # Lower expense ratio for luxury properties
        
        operating_expenses = effective_gross_income * expense_ratio
        noi = effective_gross_income - operating_expenses
        
        # Cap rate based on market and property quality
        base_cap_rate = 0.065 if quality_score >= 80 else 0.075
        if property_info.is_bridge_loan:
            base_cap_rate += 0.01  # Higher cap rate for bridge loans
        
        cap_rate = base_cap_rate * 100  # Convert to percentage
        property_value = noi / base_cap_rate
        
        # Cash return calculation
        if property_info.transaction_type == "acquisition":
            cash_return = 8.5 + (quality_score / 25)  # Higher return for acquisitions
        else:
            cash_return = 6.5 + (quality_score / 30)  # More conservative for refinance
        
        # Create demo Excel file with enhanced data - improved filename without excessive underscores
        clean_property_name = property_info.property_name.replace(' ', ' ').strip()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"{clean_property_name} Analysis {timestamp}.xlsx"
        excel_path = f"outputs/{excel_filename}"
        os.makedirs("outputs", exist_ok=True)
        
        # Create a more detailed demo Excel file
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Property Information
        summary_data = [
            ["PROPERTY ANALYSIS SUMMARY", ""],
            ["Property Name", property_info.property_name],
            ["Address", property_info.property_address],
            ["Transaction Type", property_info.transaction_type.title()],
            ["Bridge Loan", "Yes" if property_info.is_bridge_loan else "No"],
            ["Analysis Date", datetime.now().strftime("%B %d, %Y")],
            ["Estimated Units", f"{estimated_units:,}"],
            ["", ""],
            ["DOCUMENT SUMMARY", ""],
            ["Rent Roll Files", str(len(rent_roll_files))],
            ["T12/Operating Files", str(len(t12_files))],
            ["Additional Files", str(len(additional_files))],
            ["Total Documents", str(len(uploaded_files))],
            ["Data Quality Score", f"{quality_score}/100"],
            ["", ""],
            ["FINANCIAL SUMMARY", ""],
            ["Gross Potential Income", f"${gross_potential_income:,.0f}"],
            ["Vacancy Factor", f"{vacancy_factor:.1%}"],
            ["Effective Gross Income", f"${effective_gross_income:,.0f}"],
            ["Operating Expenses", f"${operating_expenses:,.0f}"],
            ["Operating Expense Ratio", f"{expense_ratio:.1%}"],
            ["Net Operating Income", f"${noi:,.0f}"],
            ["Cap Rate", f"{cap_rate:.2f}%"],
            ["Estimated Property Value", f"${property_value:,.0f}"],
            ["Cash-on-Cash Return", f"{cash_return:.2f}%"],
            ["", ""],
            ["FLAGS & NOTES", ""],
            ["Data Quality", "Excellent" if quality_score >= 80 else "Good" if quality_score >= 60 else "Limited"],
            ["Documentation", "Complete" if (rent_roll_files and t12_files) else "Partial"],
            ["Review Items", f"{3 - (len(rent_roll_files) + len(t12_files))} items need attention" if quality_score < 80 else "No major issues identified"]
        ]
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_num, column=1, value=label)
            ws_summary.cell(row=row_num, column=2, value=value)
            
            # Style headers
            if label in ["PROPERTY ANALYSIS SUMMARY", "DOCUMENT SUMMARY", "FINANCIAL SUMMARY", "FLAGS & NOTES"]:
                cell = ws_summary.cell(row=row_num, column=1)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # File Details Sheet
        ws_files = wb.create_sheet("File Details")
        file_headers = ["File Name", "Type", "Size (MB)", "Status"]
        
        for col, header in enumerate(file_headers, 1):
            cell = ws_files.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        row = 2
        for file_path, file_type in file_type_mapping.items():
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
            
            ws_files.cell(row=row, column=1, value=file_name)
            ws_files.cell(row=row, column=2, value=file_type.replace('_', ' ').title())
            ws_files.cell(row=row, column=3, value=f"{file_size:.1f}")
            ws_files.cell(row=row, column=4, value="Processed")
            row += 1
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_files]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add Rent Roll Simulation Sheet (based on actual property analysis)
        ws_rentroll = wb.create_sheet("Rent Roll Analysis")
        
        # Generate dynamic rent roll data
        rentroll_headers = ["Unit #", "Unit Type", "Sq Ft", "Current Rent", "Market Rent", "Lease Expiry", "Status"]
        for col, header in enumerate(rentroll_headers, 1):
            cell = ws_rentroll.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Generate realistic unit data
        import random
        unit_types = ["1BR/1BA", "2BR/2BA", "3BR/2BA", "Studio"]
        sq_ft_ranges = {"Studio": (450, 600), "1BR/1BA": (650, 850), "2BR/2BA": (900, 1200), "3BR/2BA": (1200, 1600)}
        
        total_rent = 0
        for i in range(1, min(estimated_units + 1, 51)):  # Limit to 50 units for demo
            unit_type = random.choice(unit_types)
            sq_ft = random.randint(*sq_ft_ranges[unit_type])
            base_rent = sq_ft * (1.5 + random.uniform(-0.2, 0.3))  # $1.30-$1.80 per sq ft
            market_rent = base_rent * random.uniform(1.02, 1.08)  # Market rent slightly higher
            
            lease_months = random.randint(1, 18)
            lease_expiry = (datetime.now() + timedelta(days=lease_months * 30)).strftime("%m/%Y")
            status = random.choice(["Occupied", "Occupied", "Occupied", "Vacant", "Notice"])
            
            if status == "Occupied":
                total_rent += base_rent
            
            row_data = [
                f"Unit {i:03d}",
                unit_type,
                sq_ft,
                f"${base_rent:,.0f}",
                f"${market_rent:,.0f}",
                lease_expiry,
                status
            ]
            
            for col, value in enumerate(row_data, 1):
                ws_rentroll.cell(row=i+1, column=col, value=value)
        
        # Add Financial Projections Sheet
        ws_projections = wb.create_sheet("Financial Projections")
        
        projections_data = [
            ["INCOME PROJECTIONS", "", "", "", ""],
            ["Line Item", "Year 1", "Year 2", "Year 3", "Notes"],
            ["Gross Potential Rent", f"${gross_potential_income:,.0f}", 
             f"${gross_potential_income * 1.03:,.0f}", 
             f"${gross_potential_income * 1.06:,.0f}", "3% annual growth"],
            ["Vacancy Loss", f"${gross_potential_income * vacancy_factor:,.0f}", 
             f"${gross_potential_income * 1.03 * vacancy_factor:,.0f}", 
             f"${gross_potential_income * 1.06 * vacancy_factor:,.0f}", f"{vacancy_factor:.1%} vacancy"],
            ["Other Income", f"${gross_potential_income * 0.02:,.0f}", 
             f"${gross_potential_income * 1.03 * 0.02:,.0f}", 
             f"${gross_potential_income * 1.06 * 0.02:,.0f}", "Parking, fees, etc."],
            ["Effective Gross Income", f"${effective_gross_income:,.0f}", 
             f"${effective_gross_income * 1.03:,.0f}", 
             f"${effective_gross_income * 1.06:,.0f}", ""],
            ["", "", "", "", ""],
            ["OPERATING EXPENSES", "", "", "", ""],
            ["Property Management", f"${effective_gross_income * 0.05:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.05:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.05:,.0f}", "5% of EGI"],
            ["Property Taxes", f"${effective_gross_income * 0.12:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.12:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.12:,.0f}", "Market rate"],
            ["Insurance", f"${effective_gross_income * 0.03:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.03:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.03:,.0f}", "Property insurance"],
            ["Maintenance & Repairs", f"${effective_gross_income * 0.08:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.08:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.08:,.0f}", "Regular maintenance"],
            ["Utilities", f"${effective_gross_income * 0.04:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.04:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.04:,.0f}", "Common areas"],
            ["Other Expenses", f"${effective_gross_income * 0.03:,.0f}", 
             f"${effective_gross_income * 1.03 * 0.03:,.0f}", 
             f"${effective_gross_income * 1.06 * 0.03:,.0f}", "Miscellaneous"],
            ["Total Operating Expenses", f"${operating_expenses:,.0f}", 
             f"${operating_expenses * 1.03:,.0f}", 
             f"${operating_expenses * 1.06:,.0f}", ""],
            ["", "", "", "", ""],
            ["NET OPERATING INCOME", f"${noi:,.0f}", 
             f"${noi * 1.03:,.0f}", 
             f"${noi * 1.06:,.0f}", ""],
            ["Cash Flow After Debt Service", f"${noi * 0.75:,.0f}", 
             f"${noi * 1.03 * 0.75:,.0f}", 
             f"${noi * 1.06 * 0.75:,.0f}", "Assuming 75% leverage"]
        ]
        
        for row_num, row_data in enumerate(projections_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_projections.cell(row=row_num, column=col_num, value=value)
                if row_num == 1 or "INCOME PROJECTIONS" in str(value) or "OPERATING EXPENSES" in str(value):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        wb.save(excel_path)
        
        # Step 7: PDF Generation
        update_progress(session_id, 7, "PDF Generation", "Creating lender-ready PDF package...")
        await asyncio.sleep(1.5)
        
        # Create demo PDF path with cleaner filename
        pdf_filename = f"{clean_property_name} Package {timestamp}.pdf"
        pdf_path = f"outputs/{pdf_filename}"
        
        # Create a proper PDF file using reportlab
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Create PDF document
        doc = SimpleDocTemplate(pdf_path, pagesize=letter, 
                              rightMargin=72, leftMargin=72, 
                              topMargin=72, bottomMargin=18)
        
        # Build PDF content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Real Estate Underwriting Analysis", title_style))
        story.append(Spacer(1, 20))
        
        # Property Information
        property_info_style = ParagraphStyle(
            'PropertyInfo',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.darkgreen
        )
        story.append(Paragraph("Property Information", property_info_style))
        
        property_data = [
            ["Property Name:", property_info.property_name],
            ["Address:", property_info.property_address],
            ["Transaction Type:", property_info.transaction_type.title()],
            ["Analysis Date:", datetime.now().strftime("%B %d, %Y")],
            ["Bridge Loan:", "Yes" if property_info.is_bridge_loan else "No"]
        ]
        
        property_table = Table(property_data, colWidths=[2*inch, 4*inch])
        property_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(property_table)
        story.append(Spacer(1, 20))
        
        # Financial Summary
        story.append(Paragraph("Financial Summary", property_info_style))
        
        financial_data = [
            ["Net Operating Income:", f"${noi:,.0f}"],
            ["Cap Rate:", f"{cap_rate:.2f}%"],
            ["Cash-on-Cash Return:", f"{cash_return:.2f}%"],
            ["Property Value:", f"${property_value:,.0f}"],
            ["Quality Score:", f"{quality_score}/100"]
        ]
        
        financial_table = Table(financial_data, colWidths=[2*inch, 4*inch])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(financial_table)
        story.append(Spacer(1, 20))
        
        # File Analysis Summary
        story.append(Paragraph("Document Analysis", property_info_style))
        
        doc_summary = f"Analysis completed on {len(uploaded_files)} uploaded documents:"
        if rent_roll_files:
            doc_summary += f"\n• {len(rent_roll_files)} Rent Roll document(s)"
        if t12_files:
            doc_summary += f"\n• {len(t12_files)} T12 Financial Statement(s)"
        if additional_files:
            doc_summary += f"\n• {len(additional_files)} Additional supporting document(s)"
        
        story.append(Paragraph(doc_summary, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Footer
        footer_text = "This analysis was generated by the Real Estate Underwriting AI System. " \
                     "For questions or additional analysis, please contact your underwriting team."
        story.append(Paragraph(footer_text, styles['Italic']))
        
        # Build PDF
        doc.build(story)
        
        # Complete processing
        session.status = "completed"
        session.current_step = 7
        session.progress_percentage = 100.0
        session.step_name = "Processing Complete"
        session.message = f"Analysis complete! Quality score: {quality_score}%"
        
        # Store results
        session.results = {
            "session_id": session_id,
            "property_info": property_info.dict(),
            "file_analysis": {
                "rent_roll_files": len(rent_roll_files),
                "t12_files": len(t12_files),
                "additional_files": len(additional_files),
                "total_files": len(uploaded_files),
                "quality_score": quality_score,
                "estimated_units": estimated_units
            },
            "underwriting_summary": {
                "noi_analysis": {
                    "net_operating_income": int(noi),
                    "effective_gross_income": int(effective_gross_income),
                    "gross_potential_income": int(gross_potential_income),
                    "operating_expenses": int(operating_expenses),
                    "expense_ratio": round(expense_ratio * 100, 1),
                    "vacancy_factor": round(vacancy_factor * 100, 1)
                },
                "valuation": {
                    "estimated_value": int(property_value),
                    "cap_rate": round(cap_rate, 2),
                    "cash_on_cash_return": round(cash_return, 2),
                    "price_per_unit": int(property_value / estimated_units) if estimated_units > 0 else 0
                }
            },
            "excel_path": excel_path,
            "pdf_path": pdf_path,
            "flags_count": max(0, 3 - len(rent_roll_files) - len(t12_files)),
            "processing_time": datetime.now().isoformat()
        }
        
        logger.info(f"✅ Demo processing completed for session {session_id} with {len(uploaded_files)} files")
        
    except Exception as e:
        logger.error(f"❌ Processing failed for session {session_id}: {str(e)}")
        session = processing_sessions[session_id]
        session.status = "error"
        session.error_message = str(e)
        session.message = f"Processing failed: {str(e)}"

def update_progress(session_id: str, step: int, step_name: str, message: str):
    """Update processing progress for a session."""
    if session_id in processing_sessions:
        session = processing_sessions[session_id]
        session.current_step = step
        session.total_steps = 7
        session.step_name = step_name
        session.progress_percentage = (step / 7) * 100
        session.message = message
        logger.info(f"📊 Session {session_id}: Step {step}/7 - {step_name}")

@app.get("/api/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

# Cleanup endpoint for development
@app.delete("/api/cleanup/{session_id}")
async def cleanup_session(session_id: str):
    """Clean up session data (development only)."""
    if session_id in processing_sessions:
        del processing_sessions[session_id]
        
        # Clean up files
        session_dir = f"uploads/{session_id}"
        if os.path.exists(session_dir):
            shutil.rmtree(session_dir)
        
        return {"message": f"Session {session_id} cleaned up"}
    
    raise HTTPException(status_code=404, detail="Session not found")

@app.post("/api/update-pdf/{session_id}")
async def update_pdf_content(
    session_id: str,
    pdf_notes: str = Form(""),
    additional_pages: List[UploadFile] = File([])
):
    """
    Update PDF package with additional notes or pages before download.
    """
    if session_id not in processing_sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session_status = processing_sessions[session_id]
    
    if session_status.status != "completed":
        raise HTTPException(status_code=400, detail="Analysis not completed yet")
    
    try:
        # Update PDF with additional content
        results = session_status.results
        if results and "pdf_path" in results:
            pdf_path = results["pdf_path"]
            
            # For demo, create an updated PDF with notes
            if pdf_notes:
                # Create updated filename
                base_name = os.path.splitext(pdf_path)[0]
                updated_pdf_path = f"{base_name}_Updated.pdf"
                
                # In a real implementation, you would modify the PDF content here
                # For demo, just copy and rename
                shutil.copy2(pdf_path, updated_pdf_path)
                
                # Update the results with new path
                session_status.results["pdf_path"] = updated_pdf_path
                session_status.results["pdf_updated"] = True
                session_status.results["update_notes"] = pdf_notes
        
        return {
            "message": "PDF updated successfully",
            "session_id": session_id,
            "pdf_path": session_status.results.get("pdf_path", ""),
            "updated": True
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update PDF: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    
    # Ensure required directories exist
    os.makedirs("uploads", exist_ok=True)
    os.makedirs("outputs", exist_ok=True)
    os.makedirs("static", exist_ok=True)
    os.makedirs("templates", exist_ok=True)
    
    print("🚀 Starting Real Estate Underwriting AI Server...")
    print("📊 Access the application at: http://localhost:8000")
    print("🎯 This is a demo version with simulated processing")
    
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
